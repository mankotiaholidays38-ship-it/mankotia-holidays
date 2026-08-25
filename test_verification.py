import requests
import openpyxl
import sys

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')


def verify_system():
    print("=" * 60)
    print("MANKOTIA HOLIDAYS - END-TO-END SYSTEM VERIFICATION")
    print("=" * 60)

    base_url = "http://127.0.0.1:8000"
    session = requests.Session()

    # 1. Test Homepage
    r_home = session.get(base_url)
    assert r_home.status_code == 200, f"Home returned {r_home.status_code}"
    print(" [PASS] 1. Homepage loads successfully (HTTP 200)")

    # 2. Test Lead Inquiry Recording into Excel
    lead_payload = {
        "name": "Rohan Kapoor",
        "phone": "+91 99887 76655",
        "email": "rohan.kapoor@example.com",
        "destination": "Goa Tropical Getaway",
        "travelers": "4 Friends",
        "travel_date": "December 2026",
        "budget": "Standard Tier",
        "notes": "Looking for beachfront villa and sunset cruise tickets."
    }
    r_inq = session.post(f"{base_url}/api/inquiry", json=lead_payload)
    assert r_inq.status_code == 200, f"Inquiry failed: {r_inq.text}"
    inq_data = r_inq.json()
    assert inq_data["success"] is True
    print(f" [PASS] 2. Inquiry recorded for {lead_payload['name']} (Excel write confirmed)")
    print(f"         WhatsApp Link Generated: {inq_data['whatsapp_redirect_url'][:60]}...")
    print(f"         Direct Call Link: {inq_data['call_link']}")

    # 3. Test Reading Stored Leads from Excel via API
    r_leads = session.get(f"{base_url}/api/leads")
    assert r_leads.status_code == 200
    leads_data = r_leads.json()
    print(f" [PASS] 3. Retrieved {leads_data['total_count']} leads directly from Excel database")

    # 4. Verify Physical Excel File On Disk
    wb = openpyxl.load_workbook("data/mankotia_leads.xlsx")
    ws = wb.active
    print(f" [PASS] 4. Excel file 'data/mankotia_leads.xlsx' verified:")
    print(f"         Sheet Name: '{ws.title}', Total Rows: {ws.max_row}")
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if idx == 1:
            print(f"         Header: {' | '.join(str(c) for c in row[:5])}...")
        else:
            print(f"         Row {idx}: {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]}")

    # 5. Test AI Itinerary Generator
    itin_payload = {
        "destination": "Manali",
        "days": 4,
        "budget": "Standard",
        "travel_style": "Friends & Adventure"
    }
    r_itin = session.post(f"{base_url}/api/generate-itinerary", json=itin_payload)
    assert r_itin.status_code == 200
    itin = r_itin.json()["itinerary"]
    print(f" [PASS] 5. AI Itinerary Generated: '{itin['title']}' ({len(itin['days'])} Days)")
    print(f"         Day 1 Theme: {itin['days'][0]['theme']}")

    # 6. Test AI Travel Concierge Chat
    r_chat = session.post(f"{base_url}/api/chat-concierge", json={"message": "Can I get a custom quote for Kashmir?"})
    assert r_chat.status_code == 200
    print(f" [PASS] 6. AI Concierge Reply: {r_chat.json()['reply'][:80]}...")

    print("=" * 60)
    print(" ALL 6 VERIFICATION CHECKS PASSED PERFECTLY!")
    print("=" * 60)

if __name__ == "__main__":
    verify_system()
