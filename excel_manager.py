import os
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEFAULT_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
EXCEL_FILE_PATH = os.path.abspath(
    os.getenv("PRIVATE_EXCEL_PATH", os.path.join(DEFAULT_DATA_DIR, "mankotia_leads.xlsx"))
)
DATA_DIR = os.path.dirname(EXCEL_FILE_PATH)
_excel_lock = threading.Lock()

PACKAGE_SHEET = "Customer Leads"
TICKET_SHEET = "Ticket Queries"
TRANSPORT_SHEET = "Transport Queries"

HEADERS = [
    "Lead ID", "Date & Time", "Customer Name", "Phone / WhatsApp", "Email Address",
    "Destination / Yatra", "Travel Date", "Pickup", "Drop", "Tour Days",
    "Number of Persons", "Children", "Children Ages", "Vehicle Category",
    "Rooms Required", "Meal Plan", "Hotel Category", "Travelers", "Budget Category",
    "Special Requirements / Notes", "Source", "Status"
]

TICKET_HEADERS = [
    "Lead ID", "Date & Time", "Customer Name", "Phone / WhatsApp", "Email Address",
    "Transit Type", "Origin (From)", "Destination (To)", "Travel Date",
    "Class / Preference", "Passengers Count", "Special Notes", "Source", "Status"
]

TRANSPORT_HEADERS = [
    "Lead ID", "Date & Time", "Customer Name", "Phone / WhatsApp", "Email Address",
    "Vehicle Category", "Rental Type", "Pickup Location", "Drop Location",
    "Pickup Date", "Duration (Days)", "Passengers Count", "Special Notes", "Source", "Status"
]

HEADER_FILL_PACKAGE = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FILL_TICKET = PatternFill(start_color="0369A1", end_color="0369A1", fill_type="solid")
HEADER_FILL_TRANSPORT = PatternFill(start_color="047857", end_color="047857", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ROW_FONT = Font(name="Calibri", size=11, color="1E293B")
BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="0F172A")

ZEBRA_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ZEBRA_FILL_EVEN = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

BORDER_LIGHT = Border(
    left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0')
)

def format_readable_time(dt_str) -> str:
    try:
        if isinstance(dt_str, datetime):
            return dt_str.strftime("%d-%b-%Y  %I:%M %p")
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%d-%b-%Y %I:%M %p"):
            try:
                return datetime.strptime(str(dt_str).strip(), fmt).strftime("%d-%b-%Y  %I:%M %p")
            except ValueError:
                continue
    except Exception:
        pass
    return str(dt_str or "")


def setup_sheet_headers(ws, headers, fill):
    ws.views.sheetView[0].showGridLines = True
    if ws.max_row < 1 or not ws.cell(row=1, column=1).value:
        ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = headers[col_num - 1]
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_LIGHT
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def ensure_excel_file_exists() -> str:
    os.makedirs(DATA_DIR, exist_ok=True)
    with _excel_lock:
        if not os.path.exists(EXCEL_FILE_PATH):
            wb = Workbook()
            ws_pkg = wb.active
            ws_pkg.title = PACKAGE_SHEET
            setup_sheet_headers(ws_pkg, HEADERS, HEADER_FILL_PACKAGE)
            
            ws_tkt = wb.create_sheet(title=TICKET_SHEET)
            setup_sheet_headers(ws_tkt, TICKET_HEADERS, HEADER_FILL_TICKET)
            
            ws_trans = wb.create_sheet(title=TRANSPORT_SHEET)
            setup_sheet_headers(ws_trans, TRANSPORT_HEADERS, HEADER_FILL_TRANSPORT)
            
            wb.save(EXCEL_FILE_PATH)
            wb.close()
        else:
            wb = load_workbook(EXCEL_FILE_PATH)
            if PACKAGE_SHEET not in wb.sheetnames:
                setup_sheet_headers(wb.create_sheet(title=PACKAGE_SHEET, index=0), HEADERS, HEADER_FILL_PACKAGE)
            if TICKET_SHEET not in wb.sheetnames:
                setup_sheet_headers(wb.create_sheet(title=TICKET_SHEET), TICKET_HEADERS, HEADER_FILL_TICKET)
            if TRANSPORT_SHEET not in wb.sheetnames:
                setup_sheet_headers(wb.create_sheet(title=TRANSPORT_SHEET), TRANSPORT_HEADERS, HEADER_FILL_TRANSPORT)
            wb.save(EXCEL_FILE_PATH)
            wb.close()
    return EXCEL_FILE_PATH


def _style_and_append(ws, row_data, centered_cols, headers_len):
    ws.append(row_data)
    current_row = ws.max_row
    ws.row_dimensions[current_row].height = 24
    row_fill = ZEBRA_FILL_EVEN if current_row % 2 == 0 else ZEBRA_FILL_ODD
    
    for col_idx in range(1, len(row_data) + 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = BOLD_FONT if col_idx in [1, 3] else ROW_FONT
        cell.border = BORDER_LIGHT
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in centered_cols else "left", vertical="center", wrap_text=True)

    for col_idx in range(1, headers_len + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(row[0] or '')) for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True)), default=10)
        ws.column_dimensions[col_letter].width = max(15, min(max_len + 4, 45))


def add_lead_to_excel(name: str, phone: str, email: str, destination: str = "General Inquiry", travel_date: str = "Flexible", pickup: str = "", drop: str = "", days: int = 1, number_of_persons: int = 1, children: int = 0, child_ages: str = "", vehicle_category: str = "No Vehicle Required", rooms_required: int = 1, meal_plan: str = "", hotel_category: str = "", travelers: str = "1-2 Travelers", budget: str = "Standard", notes: str = "", source: str = "Website Inquiry") -> dict:
    ensure_excel_file_exists()
    timestamp = datetime.now().strftime("%d-%b-%Y  %I:%M %p")
    lead_id = f"#MK-PKG-{datetime.now().strftime('%Y%m%d%H%M%S%f')[:17]}"
    row_data = [lead_id, timestamp, name.strip(), phone.strip(), email.strip(), destination.strip() or "General Inquiry", travel_date.strip() or "Flexible", pickup.strip(), drop.strip(), days, number_of_persons, children, child_ages.strip(), vehicle_category.strip() or "No Vehicle", rooms_required, meal_plan.strip(), hotel_category.strip(), travelers.strip(), budget.strip() or "Standard", notes.strip(), source.strip() or "Website", "New Inquiry"]
    
    with _excel_lock:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb[PACKAGE_SHEET] if PACKAGE_SHEET in wb.sheetnames else wb.active
        setup_sheet_headers(ws, HEADERS, HEADER_FILL_PACKAGE)
        _style_and_append(ws, row_data, [1, 2, 4, 7, 8, 9, 10, 11, 12, 15, 21, 22], len(HEADERS))
        wb.save(EXCEL_FILE_PATH)
        wb.close()
        
    return {"lead_id": lead_id, "category": "package", "timestamp": timestamp, "name": name, "phone": phone, "email": email, "destination": destination, "travel_date": travel_date}


def add_ticket_query_to_excel(name: str, phone: str, email: str, transit_type: str, origin: str, destination: str, travel_date: str, travel_class: str = "Economy", passengers: int = 1, notes: str = "", source: str = "Website Ticket Form") -> dict:
    ensure_excel_file_exists()
    timestamp = datetime.now().strftime("%d-%b-%Y  %I:%M %p")
    lead_id = f"#MK-TKT-{datetime.now().strftime('%Y%m%d%H%M%S%f')[:17]}"
    row_data = [lead_id, timestamp, name.strip(), phone.strip(), email.strip(), transit_type.strip(), origin.strip(), destination.strip(), travel_date.strip() or "Flexible", travel_class.strip(), passengers, notes.strip() or "—", source.strip(), "New Ticket Query"]
    
    with _excel_lock:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb[TICKET_SHEET] if TICKET_SHEET in wb.sheetnames else wb.create_sheet(title=TICKET_SHEET)
        setup_sheet_headers(ws, TICKET_HEADERS, HEADER_FILL_TICKET)
        _style_and_append(ws, row_data, [1, 2, 4, 6, 9, 10, 11, 13, 14], len(TICKET_HEADERS))
        wb.save(EXCEL_FILE_PATH)
        wb.close()
        
    return {"lead_id": lead_id, "category": "ticket", "timestamp": timestamp, "name": name, "phone": phone, "email": email}


def add_transport_query_to_excel(name: str, phone: str, email: str, vehicle_category: str, rental_type: str, pickup: str, drop: str, pickup_date: str, duration_days: int = 1, passengers: int = 2, notes: str = "", source: str = "Website Transport Form") -> dict:
    ensure_excel_file_exists()
    timestamp = datetime.now().strftime("%d-%b-%Y  %I:%M %p")
    lead_id = f"#MK-CAB-{datetime.now().strftime('%Y%m%d%H%M%S%f')[:17]}"
    row_data = [lead_id, timestamp, name.strip(), phone.strip(), email.strip(), vehicle_category.strip(), rental_type.strip(), pickup.strip(), drop.strip(), pickup_date.strip() or "Flexible", duration_days, passengers, notes.strip() or "—", source.strip(), "New Transport Query"]
    
    with _excel_lock:
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb[TRANSPORT_SHEET] if TRANSPORT_SHEET in wb.sheetnames else wb.create_sheet(title=TRANSPORT_SHEET)
        setup_sheet_headers(ws, TRANSPORT_HEADERS, HEADER_FILL_TRANSPORT)
        _style_and_append(ws, row_data, [1, 2, 4, 7, 10, 11, 12, 14, 15], len(TRANSPORT_HEADERS))
        wb.save(EXCEL_FILE_PATH)
        wb.close()
        
    return {"lead_id": lead_id, "category": "transport", "timestamp": timestamp, "name": name, "phone": phone, "email": email}


def get_all_leads(filter_category: str = None) -> list[dict]:
    ensure_excel_file_exists()
    all_leads = []
    
    with _excel_lock:
        wb = load_workbook(EXCEL_FILE_PATH, data_only=True)
        
        if (not filter_category or filter_category == "package") and PACKAGE_SHEET in wb.sheetnames:
            ws = wb[PACKAGE_SHEET]
            headers = [str(cell.value or "") for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    record = dict(zip(headers, row))
                    all_leads.append({
                        "lead_id": record.get("Lead ID") or "", "category": "package",
                        "category_label": "Holiday Package & Stays",
                        "timestamp": format_readable_time(record.get("Date & Time", "")),
                        "name": str(record.get("Customer Name") or "Customer").strip(),
                        "phone": str(record.get("Phone / WhatsApp") or "N/A").strip(),
                        "email": str(record.get("Email Address") or "").strip(),
                        "destination": str(record.get("Destination / Yatra") or "General Inquiry").strip(),
                        "travelers": record.get("Travelers", "") or "", "travel_date": record.get("Travel Date", "") or "",
                        "pickup": record.get("Pickup", "") or "", "drop": record.get("Drop", "") or "",
                        "days": record.get("Tour Days", "") or "", "number_of_persons": record.get("Number of Persons", "") or "",
                        "children": record.get("Children", "") or "", "child_ages": record.get("Children Ages", "") or "",
                        "vehicle_category": record.get("Vehicle Category", "") or "", "rooms_required": record.get("Rooms Required", "") or "",
                        "meal_plan": record.get("Meal Plan", "") or "", "hotel_category": record.get("Hotel Category", "") or "",
                        "budget": record.get("Budget Category", "") or "", "notes": record.get("Special Requirements / Notes", "") or "",
                        "source": record.get("Source", "Website") or "Website", "status": record.get("Status", "New") or "New"
                    })

        if (not filter_category or filter_category == "ticket") and TICKET_SHEET in wb.sheetnames:
            ws = wb[TICKET_SHEET]
            headers = [str(cell.value or "") for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    record = dict(zip(headers, row))
                    transit = record.get("Transit Type", "Ticket")
                    all_leads.append({
                        "lead_id": record.get("Lead ID") or "", "category": "ticket", "category_label": f"Ticket: {transit}",
                        "timestamp": format_readable_time(record.get("Date & Time", "")),
                        "name": record.get("Customer Name", "") or "", "phone": record.get("Phone / WhatsApp", "") or "",
                        "email": record.get("Email Address", "") or "", "destination": f"{transit}: {record.get('Origin (From)', '')} -> {record.get('Destination (To)', '')}",
                        "travelers": f"{record.get('Passengers Count', 1)} Pax", "travel_date": record.get("Travel Date", "") or "",
                        "pickup": record.get("Origin (From)", ""), "drop": record.get("Destination (To)", ""), "days": 1,
                        "notes": record.get("Special Notes", "") or "", "source": record.get("Source", "Ticket Query"), "status": record.get("Status", "New")
                    })

        if (not filter_category or filter_category == "transport") and TRANSPORT_SHEET in wb.sheetnames:
            ws = wb[TRANSPORT_SHEET]
            headers = [str(cell.value or "") for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    record = dict(zip(headers, row))
                    veh = record.get("Vehicle Category", "Cab")
                    all_leads.append({
                        "lead_id": record.get("Lead ID") or "", "category": "transport", "category_label": f"Cab: {veh}",
                        "timestamp": format_readable_time(record.get("Date & Time", "")),
                        "name": record.get("Customer Name", "") or "", "phone": record.get("Phone / WhatsApp", "") or "",
                        "email": record.get("Email Address", "") or "", "destination": f"Cab: {record.get('Pickup Location', '')} -> {record.get('Drop Location', '')}",
                        "travelers": f"{record.get('Rental Type', 'Tour')} - {veh}", "travel_date": record.get("Pickup Date", "") or "",
                        "pickup": record.get("Pickup Location", ""), "drop": record.get("Drop Location", ""), "days": record.get("Duration (Days)", 1),
                        "notes": record.get("Special Notes", "") or "", "source": record.get("Source", "Transport Query"), "status": record.get("Status", "New")
                    })

        wb.close()
    return list(reversed(all_leads))


def delete_lead(lead_id: str) -> bool:
    ensure_excel_file_exists()
    with _excel_lock:
        wb = load_workbook(EXCEL_FILE_PATH)
        for sheet_name in [PACKAGE_SHEET, TICKET_SHEET, TRANSPORT_SHEET]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [str(cell.value or "") for cell in ws[1]]
                if "Lead ID" in headers:
                    col_idx = headers.index("Lead ID") + 1
                    for r in range(2, ws.max_row + 1):
                        if str(ws.cell(row=r, column=col_idx).value or "").strip() == lead_id.strip():
                            ws.delete_rows(r, 1)
                            wb.save(EXCEL_FILE_PATH)
                            wb.close()
                            return True
        wb.close()
    return False


def delete_multiple_leads(lead_ids: list[str]) -> int:
    targets = set(str(lid).strip() for lid in lead_ids if lid)
    if not targets:
        return 0
    ensure_excel_file_exists()
    deleted_count = 0
    with _excel_lock:
        wb = load_workbook(EXCEL_FILE_PATH)
        for sheet_name in [PACKAGE_SHEET, TICKET_SHEET, TRANSPORT_SHEET]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = [str(cell.value or "") for cell in ws[1]]
                if "Lead ID" in headers:
                    col_idx = headers.index("Lead ID") + 1
                    for r in range(ws.max_row, 1, -1):
                        val = str(ws.cell(row=r, column=col_idx).value or "").strip()
                        if val in targets:
                            ws.delete_rows(r, 1)
                            deleted_count += 1
                            targets.remove(val)
        if deleted_count > 0:
            wb.save(EXCEL_FILE_PATH)
        wb.close()
    return deleted_count
